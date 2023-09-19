import os
import shutil
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook

class FileMoverApp:
    def __init__(self, master):
        self.master = master
        master.title("Pemindahan File PDF Berdasarkan Excel (Copy)")
        master.geometry("400x250")

        self.create_widgets()

    def create_widgets(self):
        frame = tk.Frame(self.master, padx=20, pady=20)
        frame.pack(expand=True, fill='both')

        excel_label = tk.Label(frame, text="Pilih File Excel:")
        excel_label.grid(row=0, column=0, sticky='w')

        self.excel_entry = tk.Entry(frame, width=40)
        self.excel_entry.grid(row=0, column=1)

        browse_excel_button = tk.Button(frame, text="Browse", command=self.browse_excel_files)
        browse_excel_button.grid(row=0, column=2)

        pdf_folder_label = tk.Label(frame, text="Pilih Folder PDF:")
        pdf_folder_label.grid(row=1, column=0, sticky='w')

        self.pdf_folder_entry = tk.Entry(frame, width=40)
        self.pdf_folder_entry.grid(row=1, column=1)

        browse_pdf_button = tk.Button(frame, text="Browse", command=self.browse_pdf_folder)
        browse_pdf_button.grid(row=1, column=2)

        copy_button = tk.Button(frame, text="Copy PDF", command=self.copy_matching_pdfs)
        copy_button.grid(row=2, column=0, columnspan=3, pady=(20, 0))

        self.result_label = tk.Label(frame, text="")
        self.result_label.grid(row=3, column=0, columnspan=3)

    def browse_excel_files(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")])
        self.excel_entry.delete(0, tk.END)
        self.excel_entry.insert(0, "\n".join(file_paths))

    def browse_pdf_folder(self):
        folder_path = filedialog.askdirectory()
        self.pdf_folder_entry.delete(0, tk.END)
        self.pdf_folder_entry.insert(0, folder_path)

    def copy_matching_pdfs(self):
        excel_files = self.excel_entry.get().split('\n')
        pdf_folder = self.pdf_folder_entry.get()

        if not excel_files or not pdf_folder:
            messagebox.showerror("Error", "Harap pilih file Excel dan folder PDF terlebih dahulu.")
            return

        try:
            for excel_file in excel_files:
                wb = load_workbook(excel_file)
                sheet = wb.active

                folder_name = os.path.splitext(os.path.basename(excel_file))[0]
                destination_folder = os.path.join(pdf_folder, folder_name)

                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
                    cell_value = row[0]
                    matching_pdf = next((pdf_file for pdf_file in os.listdir(pdf_folder) if cell_value in pdf_file), None)

                    if matching_pdf:
                        os.makedirs(destination_folder, exist_ok=True)
                        shutil.copy(os.path.join(pdf_folder, matching_pdf), os.path.join(destination_folder, matching_pdf))



                # Salin file Excel ke dalam folder
                shutil.copy(excel_file, os.path.join(destination_folder, os.path.basename(excel_file)))
                
                # Pindahkan folder hasil filtering ke direktori file Excel
                shutil.move(destination_folder, os.path.dirname(excel_file))
                
            messagebox.showinfo("Info", "Pemindahan berhasil (copy)!")
        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = FileMoverApp(root)
    root.mainloop()
